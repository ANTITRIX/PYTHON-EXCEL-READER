from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
wb=load_workbook('C:\\Users\\xcv\PycharmProjects\pythonProject1\countries.xlsx') #enter the worksheet path here
ws=wb.active
usa=wb['USA']
canada=wb['CANANDA']
sheet_max_row=ws.max_row
userchoice=int(input('enter 1 to Enter a new country to load its file or '
                     'enter 4 to exit' ))
if userchoice == 1 :
    user_second_choice=str(input('enter the country name in upper case letters'))
if userchoice == 4 :
    quit('the application has been terminated')
if user_second_choice =='EGYPT' : #here it iterate over the rows and coloumns in the sheet
    user_third_choice= int(input('enter 2 to Display the population of each state / province / governorate and total population of the country or 3 to - Display the state / province / governorate with the highest population and the lowest one'))
    if user_third_choice == 2:
        for row in range(1, 12):
            for col in range(1, 3):
                char = get_column_letter(col)
                print(ws[char + str(row)].value)
    elif user_third_choice==3 : #here it append the max and minimum population of the country
        max_pop = 0
        max_state = " "
        mini_pop =1000000000000000000000000000000000000000000
        mini_state = " "
        for row in range (1, sheet_max_row +1) :
            if max_pop < int(ws['B' + str(row)].value) :
                max_pop=ws['B'+str(row)].value
                max_state=ws['A'+str(row)].value

            if mini_pop >int( ws['B' + str(row)].value) :
                mini_pop=ws['B'+str(row)].value
                mini_state=ws['A'+str(row)].value
        print('the lowes population state is', mini_state, 'with', mini_pop, 'people')
        print('the highest population state is', max_state, 'with', max_pop, 'people')
elif user_second_choice =='USA' :
    user_third_choice= int(input('enter 2 to Display the population of each state / province / governorate and total population of the country or 3 to - Display the state / province / governorate with the highest population and the lowest one'))
    ws=usa
    if user_third_choice == 2:
        for row in range(1, 7):
            for col in range(1, 3):
                char = get_column_letter(col)
                print(ws[char + str(row)].value)
    elif user_third_choice==3 :
        ws=usa
        sheet_max_row = ws.max_row

        max_pop = 0
        max_state = " "
        mini_pop =1000000000000000000000000000000000000000000
        mini_state = " "
        for row in range (1, sheet_max_row +1) :
            if max_pop < int(ws['B' + str(row)].value) :
                max_pop=ws['B'+str(row)].value
                max_state=ws['A'+str(row)].value

            if mini_pop >int( ws['B' + str(row)].value) :
                mini_pop=ws['B'+str(row)].value
                mini_state=ws['A'+str(row)].value
        print('the lowes population state is', mini_state, 'with', mini_pop, 'people')
        print('the highest population state is', max_state, 'with', max_pop, 'people')
elif user_second_choice =='CANADA' :
    user_third_choice= int(input('enter 2 to Display the population of each state / province / governorate and total population of the country or 3 to - Display the state / province / governorate with the highest population and the lowest one'))
    ws=canada
    if user_third_choice == 2:
        for row in range(1, 7):
            for col in range(1, 3):
                char = get_column_letter(col)
                print(ws[char + str(row)].value)
    elif user_third_choice==3 :
        ws=canada
        sheet_max_row = ws.max_row

        max_pop = 0
        max_state = " "
        mini_pop =1000000000000000000000000000000000000000000
        mini_state = " "
        for row in range (1, sheet_max_row +1) :
            if max_pop < int(ws['B' + str(row)].value) :
                max_pop=ws['B'+str(row)].value
                max_state=ws['A'+str(row)].value

            if mini_pop >int( ws['B' + str(row)].value) :
                mini_pop=ws['B'+str(row)].value
                mini_state=ws['A'+str(row)].value
        print('the lowes population state is', mini_state, 'with', mini_pop, 'people')
        print('the highest population state is', max_state, 'with', max_pop, 'people')
else :
    quit('there is no sheet for this country ')